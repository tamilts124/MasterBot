import os
from pathlib import Path
from langchain_core.tools import tool
from .common import _truncate_output


def _check_dep(package: str, import_name: str = None) -> bool:
    """Check if a Python package is importable."""
    import importlib
    try:
        importlib.import_module(import_name or package)
        return True
    except ImportError:
        return False


@tool
def read_pdf(file_path: str, pages: str = "all") -> str:
    """Read and extract text from a PDF file.
    Args:
        file_path: Path to the PDF file (relative or absolute).
        pages: Page range to read. Use 'all' for the full document, or a range like '1-5' or a single page like '3'.
    """
    if not _check_dep("pypdf"):
        return "[Error] pypdf is not installed. Run: pip install pypdf"
    try:
        from pypdf import PdfReader

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"[Error] File not found: {file_path}"

        reader = PdfReader(str(path))
        total = len(reader.pages)

        # Determine page range
        if pages == "all":
            indices = list(range(total))
        elif "-" in pages:
            start, end = pages.split("-", 1)
            indices = list(range(int(start) - 1, min(int(end), total)))
        else:
            idx = int(pages) - 1
            indices = [idx] if 0 <= idx < total else []

        if not indices:
            return f"[Error] Invalid page range '{pages}' for a {total}-page document."

        parts = [f"[PDF: {path.name} | {total} pages total | Reading pages: {pages}]\n"]
        for i in indices:
            text = reader.pages[i].extract_text() or "[No text on this page]"
            parts.append(f"--- Page {i + 1} ---\n{text}")

        return _truncate_output("\n".join(parts))
    except Exception as exc:
        return f"[Error] Failed to read PDF: {exc}"


@tool
def read_docx(file_path: str) -> str:
    """Read and extract text from a Microsoft Word (.docx) file, including all paragraphs and tables.
    Args:
        file_path: Path to the .docx file (relative or absolute).
    """
    if not _check_dep("docx", "docx"):
        return "[Error] python-docx is not installed. Run: pip install python-docx"
    try:
        import docx

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"[Error] File not found: {file_path}"

        doc = docx.Document(str(path))
        parts = [f"[DOCX: {path.name}]\n"]

        for element in doc.element.body:
            tag = element.tag.split("}")[-1]
            if tag == "p":
                # Paragraph
                para = docx.text.paragraph.Paragraph(element, doc)
                text = para.text.strip()
                if text:
                    parts.append(text)
            elif tag == "tbl":
                # Table
                table = docx.table.Table(element, doc)
                parts.append("[Table]")
                for row in table.rows:
                    cells = [c.text.strip() for c in row.cells]
                    parts.append(" | ".join(cells))
                parts.append("[/Table]")

        return _truncate_output("\n".join(parts))
    except Exception as exc:
        return f"[Error] Failed to read DOCX: {exc}"


@tool
def read_excel(file_path: str, sheet: str = "all", max_rows: int = 500) -> str:
    """Read and extract data from a Microsoft Excel (.xlsx or .xls) file.
    Args:
        file_path: Path to the Excel file (relative or absolute).
        sheet: Sheet name or index (1-based) to read, or 'all' to read every sheet.
        max_rows: Maximum number of rows to read per sheet (default 500).
    """
    if not _check_dep("openpyxl"):
        return "[Error] openpyxl is not installed. Run: pip install openpyxl"
    try:
        import openpyxl

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"[Error] File not found: {file_path}"

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        # Determine which sheets to read
        if sheet == "all":
            targets = sheet_names
        elif sheet.isdigit():
            idx = int(sheet) - 1
            if idx < 0 or idx >= len(sheet_names):
                return f"[Error] Sheet index {sheet} out of range (1-{len(sheet_names)})."
            targets = [sheet_names[idx]]
        else:
            if sheet not in sheet_names:
                return f"[Error] Sheet '{sheet}' not found. Available: {sheet_names}"
            targets = [sheet]

        parts = [f"[Excel: {path.name} | Sheets: {sheet_names}]\n"]
        for name in targets:
            ws = wb[name]
            parts.append(f"=== Sheet: {name} ===")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count >= max_rows:
                    parts.append(f"[Truncated: showing first {max_rows} rows]")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                parts.append(" | ".join(cells))
                row_count += 1

        wb.close()
        return _truncate_output("\n".join(parts))
    except Exception as exc:
        return f"[Error] Failed to read Excel: {exc}"


@tool
def read_csv(file_path: str, delimiter: str = ",", max_rows: int = 500) -> str:
    """Read and display the contents of a CSV or TSV file.
    Args:
        file_path: Path to the CSV file (relative or absolute).
        delimiter: Column separator character (default ','; use '\\t' for TSV).
        max_rows: Maximum number of rows to return (default 500).
    """
    try:
        import csv

        path = Path(file_path).expanduser().resolve()
        if not path.exists():
            return f"[Error] File not found: {file_path}"

        sep = "\t" if delimiter in ("\\t", "tab") else delimiter
        parts = [f"[CSV: {path.name}]\n"]
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=sep)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    parts.append(f"[Truncated: showing first {max_rows} rows]")
                    break
                parts.append(" | ".join(row))

        return _truncate_output("\n".join(parts))
    except Exception as exc:
        return f"[Error] Failed to read CSV: {exc}"
